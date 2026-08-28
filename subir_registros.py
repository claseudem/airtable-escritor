#!/usr/bin/env python3
"""Sube registros desde un archivo Excel (o CSV) a la tabla RegistroIngresos de Airtable.

Uso basico:

    export AIRTABLE_TOKEN="patXXXXXXXXXXXXXX.XXXXXXXX..."
    python subir_registros.py registros_de_hoy.xlsx

Ver todas las opciones con:

    python subir_registros.py --help
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import os
import re
import sys
import time
import unicodedata
from pathlib import Path
from typing import Any, Iterable

try:
    import requests
except ImportError:  # pragma: no cover
    sys.exit("Falta la libreria 'requests'. Instala las dependencias con: pip install -r requirements.txt")


# ---------------------------------------------------------------------------
# Configuracion de la base de Airtable
# ---------------------------------------------------------------------------

BASE_ID = "appEItiU345wNDzPP"

# Tabla destino: RegistroIngresos
TABLA_REGISTROS = "tblWocpsI4879rdSd"
CAMPO_REGISTRO = "fldhtrcgFM7qKC7LR"      # Registro (autoNumber, solo lectura)
CAMPO_COLABORADOR = "fldS3HUjhGIucntwF"   # Colaborador (link a Colaboradores)
CAMPO_FECHA = "fldDmxOzOHZIjhkRt"         # Fecha de Visita (date)
CAMPO_NOTAS = "fldxeDiyZwmzcY5g9"         # Notas (multilineText)

# Tabla de apoyo: Colaboradores (para resolver el nombre -> record id)
TABLA_COLABORADORES = "tblg2ovqFReDW9cVu"
CAMPO_NOMBRE_COLABORADOR = "fldUhg9Qvx2f92HGb"  # Nombre (campo primario)

API_URL = "https://api.airtable.com/v0"
LOTE_MAXIMO = 10  # limite de la API de Airtable por peticion de escritura


# ---------------------------------------------------------------------------
# Alias de columnas del Excel -> campo de Airtable
#
# Si tu Excel usa otros encabezados, agregalos aqui (en minusculas y sin
# acentos; el script normaliza automaticamente antes de comparar).
# ---------------------------------------------------------------------------

ALIAS_COLUMNAS: dict[str, list[str]] = {
    "colaborador": [
        "colaborador", "colaboradora", "nombre", "nombre colaborador",
        "nombre del colaborador", "empleado", "persona", "visitante",
    ],
    "colaborador_id": [
        "colaborador id", "id colaborador", "record id", "recordid", "id airtable",
    ],
    "fecha": [
        "fecha de visita", "fecha visita", "fecha", "dia", "fecha de ingreso",
        "fecha ingreso", "fecha de registro",
    ],
    "notas": [
        "notas", "nota", "observaciones", "observacion", "comentarios",
        "comentario", "detalle",
    ],
}

# Clave interna que guarda el numero de fila real del archivo (para los mensajes).
CLAVE_FILA = "__fila__"

FORMATOS_FECHA = (
    "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y", "%Y/%m/%d",
    "%d/%m/%y", "%d-%m-%y", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M",
    "%d/%m/%Y %H:%M:%S",
)


# ---------------------------------------------------------------------------
# Utilidades
# ---------------------------------------------------------------------------

def normalizar(texto: Any) -> str:
    """Minusculas, sin acentos y sin espacios repetidos, para comparar textos."""
    if texto is None:
        return ""
    sin_acentos = unicodedata.normalize("NFKD", str(texto))
    sin_acentos = "".join(c for c in sin_acentos if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", sin_acentos).strip().lower()


def parsear_fecha(valor: Any) -> str | None:
    """Devuelve la fecha en formato ISO (YYYY-MM-DD) o None si no se pudo leer."""
    if valor is None or (isinstance(valor, str) and not valor.strip()):
        return None
    if isinstance(valor, dt.datetime):
        return valor.date().isoformat()
    if isinstance(valor, dt.date):
        return valor.isoformat()
    if isinstance(valor, (int, float)):
        # Numero de serie de Excel (base 1899-12-30).
        try:
            return (dt.date(1899, 12, 30) + dt.timedelta(days=int(valor))).isoformat()
        except (OverflowError, ValueError):
            return None
    texto = str(valor).strip()
    for formato in FORMATOS_FECHA:
        try:
            return dt.datetime.strptime(texto, formato).date().isoformat()
        except ValueError:
            continue
    return None


def limpiar(valor: Any) -> str:
    if valor is None:
        return ""
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    return str(valor).strip()


# ---------------------------------------------------------------------------
# Lectura del archivo de entrada
# ---------------------------------------------------------------------------

def leer_archivo(ruta: Path, hoja: str | None = None) -> list[dict[str, Any]]:
    """Lee un .xlsx/.xlsm o .csv y devuelve una lista de filas como diccionarios.

    Las claves son los encabezados originales de la primera fila con contenido,
    mas la clave interna CLAVE_FILA con el numero real de fila del archivo.
    """
    if ruta.suffix.lower() in {".csv", ".txt"}:
        return _leer_csv(ruta)
    if ruta.suffix.lower() in {".xlsx", ".xlsm"}:
        return _leer_excel(ruta, hoja)
    raise SystemExit(
        f"Formato no soportado: {ruta.suffix}. Usa .xlsx, .xlsm o .csv "
        "(si tienes un .xls antiguo, guardalo como .xlsx desde Excel)."
    )


def _leer_csv(ruta: Path) -> list[dict[str, Any]]:
    with ruta.open(newline="", encoding="utf-8-sig") as fh:
        muestra = fh.read(4096)
        fh.seek(0)
        try:
            dialecto = csv.Sniffer().sniff(muestra, delimiters=",;\t|")
        except csv.Error:
            dialecto = csv.excel
        filas = []
        for numero, fila in enumerate(csv.DictReader(fh, dialect=dialecto), start=2):
            registro = dict(fila)
            registro[CLAVE_FILA] = numero
            filas.append(registro)
        return filas


def _leer_excel(ruta: Path, hoja: str | None) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError:  # pragma: no cover
        raise SystemExit(
            "Falta la libreria 'openpyxl'. Instala las dependencias con: "
            "pip install -r requirements.txt"
        )

    libro = load_workbook(ruta, data_only=True, read_only=True)
    if hoja:
        if hoja not in libro.sheetnames:
            raise SystemExit(
                f"La hoja '{hoja}' no existe. Hojas disponibles: {', '.join(libro.sheetnames)}"
            )
        pagina = libro[hoja]
    else:
        pagina = libro[libro.sheetnames[0]]

    encabezados: list[str] | None = None
    resultado: list[dict[str, Any]] = []
    for numero, fila in enumerate(pagina.iter_rows(values_only=True), start=1):
        if encabezados is None:
            if fila is None or all(celda is None or str(celda).strip() == "" for celda in fila):
                continue  # saltamos filas vacias antes del encabezado
            encabezados = [limpiar(celda) for celda in fila]
            continue
        if fila is None or all(celda is None or str(celda).strip() == "" for celda in fila):
            continue  # fila vacia
        registro: dict[str, Any] = {
            encabezados[i] if i < len(encabezados) else f"columna_{i}": valor
            for i, valor in enumerate(fila)
        }
        registro[CLAVE_FILA] = numero
        resultado.append(registro)
    libro.close()
    return resultado


def mapear_columnas(encabezados: Iterable[str]) -> dict[str, str]:
    """Relaciona cada campo interno con el encabezado real del archivo."""
    disponibles = {
        normalizar(h): h for h in encabezados if limpiar(h) and h != CLAVE_FILA
    }
    mapa: dict[str, str] = {}
    for campo, alias in ALIAS_COLUMNAS.items():
        for nombre in alias:
            if nombre in disponibles:
                mapa[campo] = disponibles[nombre]
                break
    return mapa


# ---------------------------------------------------------------------------
# Cliente de Airtable
# ---------------------------------------------------------------------------

class Airtable:
    def __init__(self, token: str, base_id: str = BASE_ID, tiempo_espera: int = 30):
        self.sesion = requests.Session()
        self.sesion.headers.update({
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json",
        })
        self.base_id = base_id
        self.tiempo_espera = tiempo_espera

    def _peticion(self, metodo: str, tabla: str, **kwargs) -> dict:
        url = f"{API_URL}/{self.base_id}/{tabla}"
        espera = 2.0
        for intento in range(5):
            try:
                respuesta = self.sesion.request(
                    metodo, url, timeout=self.tiempo_espera, **kwargs
                )
            except requests.RequestException as exc:
                if intento == 4:
                    raise SystemExit(f"Error de red hablando con Airtable: {exc}")
                time.sleep(espera)
                espera *= 2
                continue

            if respuesta.status_code == 429 or respuesta.status_code >= 500:
                if intento == 4:
                    raise SystemExit(
                        f"Airtable respondio {respuesta.status_code}: {respuesta.text[:500]}"
                    )
                time.sleep(espera)
                espera *= 2
                continue

            if respuesta.status_code == 401:
                raise SystemExit(
                    "Airtable rechazo el token (401). Revisa AIRTABLE_TOKEN y que el "
                    "personal access token tenga acceso a esta base."
                )
            if respuesta.status_code == 403:
                raise SystemExit(
                    "Sin permisos (403). El token necesita los scopes "
                    "data.records:read y data.records:write sobre esta base."
                )
            if not respuesta.ok:
                raise SystemExit(
                    f"Airtable respondio {respuesta.status_code}: {respuesta.text[:500]}"
                )
            return respuesta.json()
        raise SystemExit("No se pudo contactar a Airtable tras varios intentos.")

    def listar(self, tabla: str, campos: list[str] | None = None) -> list[dict]:
        registros: list[dict] = []
        params: list[tuple[str, str]] = [
            ("pageSize", "100"),
            ("returnFieldsByFieldId", "true"),
        ]
        if campos:
            params += [("fields[]", campo) for campo in campos]
        offset: str | None = None
        while True:
            consulta = list(params)
            if offset:
                consulta.append(("offset", offset))
            datos = self._peticion("GET", tabla, params=consulta)
            registros.extend(datos.get("records", []))
            offset = datos.get("offset")
            if not offset:
                return registros

    def crear(self, tabla: str, registros: list[dict], typecast: bool = True) -> list[dict]:
        creados: list[dict] = []
        for inicio in range(0, len(registros), LOTE_MAXIMO):
            lote = registros[inicio:inicio + LOTE_MAXIMO]
            datos = self._peticion(
                "POST", tabla, json={"records": lote, "typecast": typecast}
            )
            creados.extend(datos.get("records", []))
            if inicio + LOTE_MAXIMO < len(registros):
                time.sleep(0.25)  # nos mantenemos por debajo de 5 req/seg
        return creados


# ---------------------------------------------------------------------------
# Logica principal
# ---------------------------------------------------------------------------

def cargar_colaboradores(api: Airtable) -> dict[str, tuple[str, str]]:
    """Devuelve {nombre normalizado: (record_id, nombre real)}."""
    directorio: dict[str, tuple[str, str]] = {}
    for registro in api.listar(TABLA_COLABORADORES, campos=[CAMPO_NOMBRE_COLABORADOR]):
        nombre = limpiar(registro.get("fields", {}).get(CAMPO_NOMBRE_COLABORADOR))
        if nombre:
            directorio[normalizar(nombre)] = (registro["id"], nombre)
    return directorio


def cargar_existentes(api: Airtable) -> set[tuple[str, str]]:
    """Pares (record id del colaborador, fecha ISO) que ya estan en RegistroIngresos."""
    existentes: set[tuple[str, str]] = set()
    for registro in api.listar(TABLA_REGISTROS, campos=[CAMPO_COLABORADOR, CAMPO_FECHA]):
        campos = registro.get("fields", {})
        fecha = limpiar(campos.get(CAMPO_FECHA))[:10]
        for enlace in campos.get(CAMPO_COLABORADOR) or []:
            id_colaborador = enlace if isinstance(enlace, str) else enlace.get("id", "")
            if id_colaborador and fecha:
                existentes.add((id_colaborador, fecha))
    return existentes


def construir_registros(
    filas: list[dict[str, Any]],
    mapa: dict[str, str],
    directorio: dict[str, tuple[str, str]],
    fecha_por_defecto: str | None,
    crear_faltantes: bool,
) -> tuple[list[dict], list[str], list[str], dict[str, str]]:
    """Convierte las filas del Excel en payloads de Airtable.

    Devuelve (registros, errores, nombres_nuevos, nombre_normalizado -> nombre real).
    """
    registros: list[dict] = []
    errores: list[str] = []
    nuevos: dict[str, str] = {}

    for fila in filas:
        indice = fila.get(CLAVE_FILA, "?")
        nombre = limpiar(fila.get(mapa.get("colaborador", ""), ""))
        id_directo = limpiar(fila.get(mapa.get("colaborador_id", ""), ""))
        fecha_bruta = fila.get(mapa.get("fecha", ""), None)
        notas = limpiar(fila.get(mapa.get("notas", ""), ""))

        if not nombre and not id_directo:
            errores.append(f"Fila {indice}: falta el colaborador.")
            continue

        fecha = parsear_fecha(fecha_bruta) or fecha_por_defecto
        if not fecha:
            errores.append(
                f"Fila {indice} ({nombre or id_directo}): fecha vacia o ilegible "
                f"({fecha_bruta!r}). Usa YYYY-MM-DD o DD/MM/AAAA, o pasa --fecha."
            )
            continue

        if id_directo.startswith("rec") and len(id_directo) == 17:
            id_colaborador: str | None = id_directo
        else:
            clave = normalizar(nombre)
            encontrado = directorio.get(clave)
            if encontrado:
                id_colaborador = encontrado[0]
            elif crear_faltantes:
                id_colaborador = None  # se resuelve despues de crearlo
                nuevos[clave] = nombre
            else:
                errores.append(
                    f"Fila {indice}: '{nombre}' no existe en Colaboradores. "
                    "Corrige el nombre o usa --crear-colaboradores."
                )
                continue

        registros.append({
            "_fila": indice,
            "_nombre": nombre,
            "_clave": normalizar(nombre),
            "_id_colaborador": id_colaborador,
            "_fecha": fecha,
            "_notas": notas,
        })

    return registros, errores, sorted(set(nuevos.values())), nuevos


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Sube registros de asistencia desde Excel a la tabla "
                    "RegistroIngresos de Airtable.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "Columnas esperadas en el Excel:\n"
            "  Colaborador       nombre tal como aparece en la tabla Colaboradores\n"
            "  Fecha de Visita   fecha del ingreso (YYYY-MM-DD o DD/MM/AAAA)\n"
            "  Notas             texto libre (opcional)\n"
            "\nOpcionalmente puedes usar una columna 'Colaborador ID' con el recId.\n"
        ),
    )
    parser.add_argument("archivo", nargs="?", help="Ruta al .xlsx, .xlsm o .csv con los registros.")
    parser.add_argument("--hoja", help="Nombre de la hoja del Excel (por defecto, la primera).")
    parser.add_argument(
        "--fecha",
        help="Fecha por defecto (YYYY-MM-DD) para las filas sin fecha. "
             "Usa 'hoy' para la fecha actual.",
    )
    parser.add_argument(
        "--crear-colaboradores", action="store_true",
        help="Crea en la tabla Colaboradores los nombres que no existan todavia.",
    )
    parser.add_argument(
        "--permitir-duplicados", action="store_true",
        help="No omite filas que ya existen en Airtable con el mismo colaborador y fecha.",
    )
    parser.add_argument(
        "--simulacion", "--dry-run", dest="simulacion", action="store_true",
        help="Muestra lo que se subiria sin escribir nada en Airtable.",
    )
    parser.add_argument(
        "--token", help="Token de Airtable (por defecto se lee de AIRTABLE_TOKEN).",
    )
    parser.add_argument(
        "--crear-plantilla", metavar="RUTA",
        help="Genera un Excel vacio con las columnas correctas y termina.",
    )
    args = parser.parse_args()

    if args.crear_plantilla:
        crear_plantilla(Path(args.crear_plantilla))
        return 0

    if not args.archivo:
        parser.error("indica el archivo de Excel a subir (o usa --crear-plantilla).")

    ruta = Path(args.archivo).expanduser()
    if not ruta.is_file():
        print(f"No encuentro el archivo: {ruta}", file=sys.stderr)
        return 1

    fecha_por_defecto = None
    if args.fecha:
        fecha_por_defecto = (
            dt.date.today().isoformat() if args.fecha.lower() in {"hoy", "today"}
            else parsear_fecha(args.fecha)
        )
        if not fecha_por_defecto:
            print(f"--fecha no es una fecha valida: {args.fecha}", file=sys.stderr)
            return 1

    token = args.token or os.environ.get("AIRTABLE_TOKEN") or os.environ.get("AIRTABLE_API_KEY")
    if not token:
        print(
            "Falta el token de Airtable. Exporta AIRTABLE_TOKEN o pasa --token.\n"
            "Crea uno en https://airtable.com/create/tokens con los scopes "
            "data.records:read y data.records:write sobre esta base.",
            file=sys.stderr,
        )
        return 1

    filas = leer_archivo(ruta, args.hoja)
    if not filas:
        print("El archivo no tiene filas de datos.", file=sys.stderr)
        return 1

    mapa = mapear_columnas(filas[0].keys())
    if "colaborador" not in mapa and "colaborador_id" not in mapa:
        print(
            "No encuentro una columna de colaborador en el Excel.\n"
            f"Encabezados leidos: {', '.join(str(k) for k in filas[0] if k != CLAVE_FILA)}\n"
            "Renombra la columna a 'Colaborador' o agrega el alias en ALIAS_COLUMNAS.",
            file=sys.stderr,
        )
        return 1
    if "fecha" not in mapa and not fecha_por_defecto:
        print(
            "No encuentro una columna de fecha. Renombrala a 'Fecha de Visita' "
            "o pasa --fecha hoy.",
            file=sys.stderr,
        )
        return 1

    print(f"Leidas {len(filas)} filas de {ruta.name}")
    print("Columnas detectadas: " + ", ".join(f"{k} -> '{v}'" for k, v in mapa.items()))

    api = Airtable(token)
    directorio = cargar_colaboradores(api)
    print(f"Colaboradores en Airtable: {len(directorio)}")

    registros, errores, nombres_nuevos, _ = construir_registros(
        filas, mapa, directorio, fecha_por_defecto, args.crear_colaboradores
    )

    # Alta de colaboradores nuevos (si se pidio).
    if nombres_nuevos:
        if args.simulacion:
            print(f"[simulacion] Se crearian {len(nombres_nuevos)} colaboradores nuevos: "
                  + ", ".join(nombres_nuevos))
        else:
            creados = api.crear(
                TABLA_COLABORADORES,
                [{"fields": {CAMPO_NOMBRE_COLABORADOR: nombre}} for nombre in nombres_nuevos],
            )
            # Airtable devuelve los registros creados en el mismo orden en que se enviaron.
            for nombre, registro in zip(nombres_nuevos, creados):
                directorio[normalizar(nombre)] = (registro["id"], nombre)
            print(f"Colaboradores creados: {len(creados)} ({', '.join(nombres_nuevos)})")
        for registro in registros:
            if registro["_id_colaborador"] is None:
                encontrado = directorio.get(registro["_clave"])
                registro["_id_colaborador"] = encontrado[0] if encontrado else "recSIMULADO"

    # Omitimos duplicados ya presentes en Airtable.
    omitidos = 0
    if not args.permitir_duplicados:
        existentes = cargar_existentes(api)
        vistos: set[tuple[str, str]] = set()
        filtrados = []
        for registro in registros:
            clave = (registro["_id_colaborador"], registro["_fecha"])
            if clave in existentes or clave in vistos:
                omitidos += 1
                continue
            vistos.add(clave)
            filtrados.append(registro)
        registros = filtrados

    payload = [
        {
            "fields": {
                CAMPO_COLABORADOR: [registro["_id_colaborador"]],
                CAMPO_FECHA: registro["_fecha"],
                **({CAMPO_NOTAS: registro["_notas"]} if registro["_notas"] else {}),
            }
        }
        for registro in registros
    ]

    print("-" * 60)
    for registro in registros[:10]:
        print(f"  fila {str(registro['_fila']):>4}  {registro['_fecha']}  "
              f"{registro['_nombre'] or registro['_id_colaborador']}  {registro['_notas'][:40]}")
    if len(registros) > 10:
        print(f"  ... y {len(registros) - 10} filas mas")
    print("-" * 60)

    if errores:
        print(f"\nFilas con problemas ({len(errores)}):", file=sys.stderr)
        for error in errores:
            print(f"  - {error}", file=sys.stderr)

    if omitidos:
        print(f"\nOmitidas {omitidos} filas que ya existian en Airtable (mismo colaborador y fecha).")

    if not payload:
        print("\nNo hay nada nuevo que subir.")
        return 1 if errores else 0

    if args.simulacion:
        print(f"\n[simulacion] Se subirian {len(payload)} registros. No se escribio nada.")
        return 1 if errores else 0

    creados = api.crear(TABLA_REGISTROS, payload)
    print(f"\nListo: {len(creados)} registros creados en RegistroIngresos.")
    print(f"Ver la tabla: https://airtable.com/{BASE_ID}/{TABLA_REGISTROS}")
    return 1 if errores else 0


def crear_plantilla(ruta: Path) -> None:
    try:
        from openpyxl import Workbook
    except ImportError:  # pragma: no cover
        raise SystemExit("Falta 'openpyxl'. Instala con: pip install -r requirements.txt")

    libro = Workbook()
    hoja = libro.active
    hoja.title = "Registros"
    hoja.append(["Colaborador", "Fecha de Visita", "Notas"])
    hoja.append(["Juan Carlos Perez", dt.date.today().isoformat(), "Ingreso puntual"])
    for columna, ancho in zip("ABC", (30, 18, 50)):
        hoja.column_dimensions[columna].width = ancho
    libro.save(ruta)
    print(f"Plantilla creada en {ruta}")


if __name__ == "__main__":
    raise SystemExit(main())
